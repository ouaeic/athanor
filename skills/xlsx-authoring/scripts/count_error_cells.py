#!/usr/local/lib/athanor/python/bin/python3
"""Count Excel error cells in a recalculated workbook and report them as JSON.

Run this against the copy produced by ``athanor-office-convert model.xlsx
model.recalc.xlsx``, not against the workbook
openpyxl just wrote: openpyxl stores formula strings without cached results, so a freshly written
workbook has no values to check and would always report zero.

Usage:
    count_error_cells.py WORKBOOK [--max-report N]

Exit status is 0 when no error cell is present and 1 otherwise, so it can gate a verify step.
"""

from __future__ import annotations

import argparse
import json
import sys

ERROR_VALUES = (
    "#REF!",
    "#VALUE!",
    "#NAME?",
    "#DIV/0!",
    "#N/A",
    "#NULL!",
    "#NUM!",
    "#SPILL!",
    "#CALC!",
)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook")
    parser.add_argument("--max-report", type=int, default=25)
    arguments = parser.parse_args()

    try:
        from openpyxl import load_workbook
    except ImportError:
        print(json.dumps({"status": "error", "reason": "openpyxl is not installed"}))
        return 2

    try:
        workbook = load_workbook(arguments.workbook, data_only=True)
    except Exception as error:  # noqa: BLE001 - the reason is reported, not swallowed
        print(json.dumps({"status": "error", "reason": str(error)}))
        return 2

    errors = []
    uncached = 0
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                if isinstance(value, str):
                    if value in ERROR_VALUES:
                        errors.append(
                            {"sheet": sheet.title, "cell": cell.coordinate, "value": value}
                        )
                    elif value.startswith("="):
                        # A formula string in a data_only read means no cached result exists,
                        # which is exactly the "looks fine, shows blank on the owner's machine"
                        # failure this check is here to catch.
                        uncached += 1

    payload = {
        "status": "success" if not errors and not uncached else "failed",
        "workbook": arguments.workbook,
        "error_cells": len(errors),
        "uncached_formulas": uncached,
        "errors": errors[: arguments.max_report],
    }
    print(json.dumps(payload, indent=2))
    return 0 if payload["status"] == "success" else 1


if __name__ == "__main__":
    sys.exit(main())
